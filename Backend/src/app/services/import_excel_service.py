"""Parse Excel/CSV uploads into import job row dicts — P3."""

from __future__ import annotations

import csv
import io
from typing import Any

from app.core.exceptions import ValidationAppError


def parse_upload(*, filename: str, content: bytes) -> list[dict[str, Any]]:
    """Return row dicts from the first sheet or CSV file."""

    lower = filename.lower()
    if lower.endswith(".csv"):
        return _parse_csv(content)
    if lower.endswith((".xlsx", ".xlsm")):
        return _parse_xlsx(content)
    raise ValidationAppError("Unsupported file type. Use .csv or .xlsx")


def _parse_csv(content: bytes) -> list[dict[str, Any]]:
    text = content.decode("utf-8-sig", errors="replace")
    reader = csv.DictReader(io.StringIO(text))
    return [dict(row) for row in reader]


def _parse_xlsx(content: bytes) -> list[dict[str, Any]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise ValidationAppError(
            "Excel support requires openpyxl. Install with: pip install openpyxl"
        ) from exc

    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    sheet = wb.active
    rows_iter = sheet.iter_rows(values_only=True)
    try:
        header_row = next(rows_iter)
    except StopIteration:
        return []
    headers = [
        str(h).strip() if h is not None else f"col{i}"
        for i, h in enumerate(header_row)
    ]
    out: list[dict[str, Any]] = []
    for row in rows_iter:
        if row is None or all(c is None or str(c).strip() == "" for c in row):
            continue
        item: dict[str, Any] = {}
        for i, key in enumerate(headers):
            if i < len(row):
                val = row[i]
                item[key] = val if val is not None else ""
        out.append(item)
    wb.close()
    return out
